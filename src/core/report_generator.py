"""Report generator for Excel, PDF, and advanced exports."""

from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any

from src.core.cost_calculator import CostCalculator, CostConfig
from src.core.dashboard_service import DashboardService, format_currency, format_size
from src.data.models import InactiveMailbox
from src.utils.logging import get_logger

if TYPE_CHECKING:
    from src.data.session import SessionManager

logger = get_logger(__name__)


class ReportFormat(Enum):
    """Supported report formats."""

    EXCEL = "xlsx"
    PDF = "pdf"
    CSV = "csv"
    JSON = "json"
    HTML = "html"


class ReportType(Enum):
    """Types of reports."""

    FULL_INVENTORY = "full_inventory"
    COST_SUMMARY = "cost_summary"
    HOLD_ANALYSIS = "hold_analysis"
    EXECUTIVE_SUMMARY = "executive_summary"
    AUDIT_LOG = "audit_log"


@dataclass
class ReportConfig:
    """Configuration for report generation."""

    format: ReportFormat = ReportFormat.EXCEL
    report_type: ReportType = ReportType.FULL_INVENTORY
    include_charts: bool = True
    include_recommendations: bool = True
    include_raw_data: bool = False
    date_range: tuple[datetime, datetime] | None = None
    filters: dict[str, Any] | None = None
    title: str | None = None
    author: str | None = None


@dataclass
class ReportMetadata:
    """Metadata for generated reports."""

    title: str
    report_type: ReportType
    generated_at: datetime
    generated_by: str
    tenant_id: str | None = None
    record_count: int = 0
    filters_applied: dict[str, Any] = field(default_factory=dict)
    version: str = "1.0"

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary."""
        return {
            "title": self.title,
            "report_type": self.report_type.value,
            "generated_at": self.generated_at.isoformat(),
            "generated_by": self.generated_by,
            "tenant_id": self.tenant_id,
            "record_count": self.record_count,
            "filters_applied": self.filters_applied,
            "version": self.version,
        }


@dataclass
class ReportResult:
    """Result of report generation."""

    success: bool
    file_path: Path | None = None
    metadata: ReportMetadata | None = None
    error: str | None = None
    file_size_bytes: int = 0
    generation_time_seconds: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary."""
        return {
            "success": self.success,
            "file_path": str(self.file_path) if self.file_path else None,
            "metadata": self.metadata.to_dict() if self.metadata else None,
            "error": self.error,
            "file_size_bytes": self.file_size_bytes,
            "generation_time_seconds": self.generation_time_seconds,
        }


class ReportGeneratorError(Exception):
    """Raised when report generation fails."""

    pass


class ExcelReportGenerator:
    """Generate Excel reports with multiple sheets and formatting.

    Uses openpyxl for Excel generation with support for:
    - Multiple worksheets
    - Formatted headers and data
    - Auto-filters and column widths
    - Basic charts (if available)
    """

    def __init__(
        self,
        session: "SessionManager",
        cost_config: CostConfig | None = None,
    ) -> None:
        """Initialize Excel generator.

        Args:
            session: Session manager
            cost_config: Optional cost configuration
        """
        self._session = session
        self._db = session.db
        self._cost_calculator = CostCalculator(session, cost_config)
        self._dashboard = DashboardService(session, cost_config)

        logger.debug("ExcelReportGenerator initialized")

    def generate_full_report(
        self,
        path: Path,
        config: ReportConfig,
    ) -> ReportResult:
        """Generate full inventory Excel report.

        Args:
            path: Output file path
            config: Report configuration

        Returns:
            ReportResult with status
        """
        start_time = datetime.now()

        try:
            # Import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                return ReportResult(
                    success=False,
                    error="openpyxl is required for Excel reports. Install with: pip install openpyxl",
                )

            # Get data
            mailboxes = self._db.get_all_mailboxes()
            cost_report = self._cost_calculator.generate_cost_report(mailboxes)
            dashboard = self._dashboard.generate_dashboard(mailboxes)

            # Create workbook
            wb = Workbook()

            # Create sheets
            self._create_summary_sheet(wb, cost_report, dashboard)
            self._create_mailbox_sheet(wb, mailboxes)
            self._create_cost_sheet(wb, cost_report)
            self._create_hold_sheet(wb, mailboxes)

            # Remove default sheet if we created others
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            # Save workbook
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)

            # Get file size
            file_size = path.stat().st_size

            # Create metadata
            metadata = ReportMetadata(
                title=config.title or "Inactive Mailbox Report",
                report_type=config.report_type,
                generated_at=datetime.now(),
                generated_by=config.author or "System",
                record_count=len(mailboxes),
                filters_applied=config.filters or {},
            )

            logger.info(f"Excel report generated: {path}")

            return ReportResult(
                success=True,
                file_path=path,
                metadata=metadata,
                file_size_bytes=file_size,
                generation_time_seconds=(datetime.now() - start_time).total_seconds(),
            )

        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            return ReportResult(
                success=False,
                error=str(e),
                generation_time_seconds=(datetime.now() - start_time).total_seconds(),
            )

    def _create_summary_sheet(self, wb, cost_report, dashboard) -> None:
        """Create summary worksheet."""
        from openpyxl.styles import Font, PatternFill, Alignment

        ws = wb.active
        ws.title = "Summary"

        # Header style
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Title
        ws["A1"] = "Inactive Mailbox Summary Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Generation info
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A3"] = ""

        # Key metrics
        row = 5
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = header_font

        metrics = [
            ("Total Inactive Mailboxes", f"{cost_report.summary.total_mailboxes:,}"),
            ("Total Monthly Cost", format_currency(cost_report.summary.total_monthly_cost)),
            ("Total Annual Cost", format_currency(cost_report.summary.total_annual_cost)),
            ("Potential Monthly Savings", format_currency(cost_report.summary.potential_savings)),
            ("Average Monthly Cost", format_currency(cost_report.summary.average_monthly_cost)),
            ("Average Age (days)", f"{cost_report.summary.average_age_days:.0f}"),
        ]

        for i, (label, value) in enumerate(metrics):
            ws[f"A{row + 1 + i}"] = label
            ws[f"B{row + 1 + i}"] = value

        # Cost by license type
        row += len(metrics) + 3
        ws[f"A{row}"] = "Cost by License Type"
        ws[f"A{row}"].font = header_font

        for i, (license_type, cost) in enumerate(cost_report.summary.by_license_type.items()):
            ws[f"A{row + 1 + i}"] = license_type
            ws[f"B{row + 1 + i}"] = format_currency(cost)

        # Recommendations
        row += len(cost_report.summary.by_license_type) + 3
        ws[f"A{row}"] = "Recommendations"
        ws[f"A{row}"].font = header_font

        for i, rec in enumerate(cost_report.recommendations):
            ws[f"A{row + 1 + i}"] = rec
            ws.merge_cells(f"A{row + 1 + i}:D{row + 1 + i}")

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

    def _create_mailbox_sheet(self, wb, mailboxes: list[InactiveMailbox]) -> None:
        """Create mailbox inventory worksheet."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Mailboxes")

        # Headers
        headers = [
            "Display Name",
            "Primary SMTP",
            "Exchange GUID",
            "Size (MB)",
            "Item Count",
            "Disconnected Date",
            "Litigation Hold",
            "In-Place Holds",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row_num, mailbox in enumerate(mailboxes, start=2):
            ws.cell(row=row_num, column=1, value=mailbox.display_name)
            ws.cell(row=row_num, column=2, value=mailbox.primary_smtp)
            ws.cell(row=row_num, column=3, value=mailbox.identity)
            ws.cell(row=row_num, column=4, value=round(mailbox.total_item_size_mb, 2))
            ws.cell(row=row_num, column=5, value=mailbox.total_item_count)
            ws.cell(row=row_num, column=6, value=mailbox.disconnected_date.strftime("%Y-%m-%d") if mailbox.disconnected_date else "")
            ws.cell(row=row_num, column=7, value="Yes" if mailbox.litigation_hold_enabled else "No")
            ws.cell(row=row_num, column=8, value=", ".join(mailbox.in_place_holds) if mailbox.in_place_holds else "None")

        # Auto-filter
        ws.auto_filter.ref = f"A1:H{len(mailboxes) + 1}"

        # Column widths
        column_widths = [30, 35, 40, 12, 12, 15, 12, 40]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_cost_sheet(self, wb, cost_report) -> None:
        """Create cost analysis worksheet."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Cost Analysis")

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Top cost mailboxes
        ws["A1"] = "Top Cost Mailboxes"
        ws["A1"].font = Font(bold=True, size=12)

        headers = ["Display Name", "License Type", "Monthly Cost", "Age (days)", "Size"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_num, info in enumerate(cost_report.top_cost_mailboxes, start=3):
            ws.cell(row=row_num, column=1, value=info.display_name)
            ws.cell(row=row_num, column=2, value=info.license_type.value)
            ws.cell(row=row_num, column=3, value=format_currency(info.monthly_cost))
            ws.cell(row=row_num, column=4, value=info.age_days)
            ws.cell(row=row_num, column=5, value=format_size(info.size_mb))

        # Column widths
        column_widths = [30, 25, 15, 12, 15]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_hold_sheet(self, wb, mailboxes: list[InactiveMailbox]) -> None:
        """Create hold analysis worksheet."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet("Hold Analysis")

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Count holds
        hold_counts: dict[str, int] = {}
        for mailbox in mailboxes:
            if mailbox.litigation_hold_enabled:
                hold_counts["Litigation Hold"] = hold_counts.get("Litigation Hold", 0) + 1
            if mailbox.in_place_holds:
                for hold in mailbox.in_place_holds:
                    if hold.startswith("UniH"):
                        hold_counts["eDiscovery Hold"] = hold_counts.get("eDiscovery Hold", 0) + 1
                    elif hold.startswith("mbx"):
                        hold_counts["In-Place Hold"] = hold_counts.get("In-Place Hold", 0) + 1
                    elif hold.startswith("cld"):
                        hold_counts["Retention Policy"] = hold_counts.get("Retention Policy", 0) + 1
                    else:
                        hold_counts["Other Hold"] = hold_counts.get("Other Hold", 0) + 1
            if not mailbox.litigation_hold_enabled and not mailbox.in_place_holds:
                hold_counts["No Hold"] = hold_counts.get("No Hold", 0) + 1

        # Title
        ws["A1"] = "Hold Type Distribution"
        ws["A1"].font = Font(bold=True, size=12)

        # Headers
        headers = ["Hold Type", "Count", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        total = sum(hold_counts.values())
        for row_num, (hold_type, count) in enumerate(sorted(hold_counts.items(), key=lambda x: x[1], reverse=True), start=3):
            ws.cell(row=row_num, column=1, value=hold_type)
            ws.cell(row=row_num, column=2, value=count)
            ws.cell(row=row_num, column=3, value=f"{count / total * 100:.1f}%")

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12


class HTMLReportGenerator:
    """Generate HTML reports for web viewing or PDF conversion."""

    def __init__(
        self,
        session: "SessionManager",
        cost_config: CostConfig | None = None,
    ) -> None:
        """Initialize HTML generator."""
        self._session = session
        self._db = session.db
        self._cost_calculator = CostCalculator(session, cost_config)
        self._dashboard = DashboardService(session, cost_config)

        logger.debug("HTMLReportGenerator initialized")

    def generate_report(
        self,
        path: Path,
        config: ReportConfig,
    ) -> ReportResult:
        """Generate HTML report.

        Args:
            path: Output file path
            config: Report configuration

        Returns:
            ReportResult with status
        """
        start_time = datetime.now()

        try:
            mailboxes = self._db.get_all_mailboxes()
            cost_report = self._cost_calculator.generate_cost_report(mailboxes)
            dashboard = self._dashboard.generate_dashboard(mailboxes)

            html = self._build_html(cost_report, dashboard, config)

            path.parent.mkdir(parents=True, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)

            metadata = ReportMetadata(
                title=config.title or "Inactive Mailbox Report",
                report_type=config.report_type,
                generated_at=datetime.now(),
                generated_by=config.author or "System",
                record_count=len(mailboxes),
            )

            return ReportResult(
                success=True,
                file_path=path,
                metadata=metadata,
                file_size_bytes=path.stat().st_size,
                generation_time_seconds=(datetime.now() - start_time).total_seconds(),
            )

        except Exception as e:
            logger.error(f"HTML generation failed: {e}")
            return ReportResult(
                success=False,
                error=str(e),
                generation_time_seconds=(datetime.now() - start_time).total_seconds(),
            )

    def _build_html(self, cost_report, dashboard, config) -> str:
        """Build HTML content."""
        title = config.title or "Inactive Mailbox Report"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
        .metric-card {{ background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }}
        .metric-value {{ font-size: 28px; font-weight: bold; color: #2c3e50; }}
        .metric-label {{ color: #7f8c8d; font-size: 14px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #3498db; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #ecf0f1; }}
        tr:hover {{ background: #f8f9fa; }}
        .recommendation {{ background: #e8f6f3; border-left: 4px solid #27ae60; padding: 15px; margin: 10px 0; }}
        .footer {{ text-align: center; color: #95a5a6; margin-top: 40px; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

        <h2>Executive Summary</h2>
        <div class="metrics">
            <div class="metric-card">
                <div class="metric-value">{cost_report.summary.total_mailboxes:,}</div>
                <div class="metric-label">Total Mailboxes</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{format_currency(cost_report.summary.total_monthly_cost)}</div>
                <div class="metric-label">Monthly Cost</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{format_currency(cost_report.summary.total_annual_cost)}</div>
                <div class="metric-label">Annual Cost</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">{format_currency(cost_report.summary.potential_savings)}</div>
                <div class="metric-label">Potential Savings</div>
            </div>
        </div>

        <h2>Cost by License Type</h2>
        <table>
            <tr><th>License Type</th><th>Monthly Cost</th><th>Percentage</th></tr>
"""
        total_cost = cost_report.summary.total_monthly_cost
        for license_type, cost in sorted(cost_report.summary.by_license_type.items(), key=lambda x: x[1], reverse=True):
            pct = (cost / total_cost * 100) if total_cost > 0 else 0
            html += f"            <tr><td>{license_type}</td><td>{format_currency(cost)}</td><td>{pct:.1f}%</td></tr>\n"

        html += """        </table>

        <h2>Top Cost Mailboxes</h2>
        <table>
            <tr><th>Display Name</th><th>License Type</th><th>Monthly Cost</th><th>Age (days)</th></tr>
"""
        for info in cost_report.top_cost_mailboxes[:10]:
            html += f"            <tr><td>{info.display_name}</td><td>{info.license_type.value}</td><td>{format_currency(info.monthly_cost)}</td><td>{info.age_days}</td></tr>\n"

        html += """        </table>

        <h2>Recommendations</h2>
"""
        for rec in cost_report.recommendations:
            html += f'        <div class="recommendation">{rec}</div>\n'

        html += f"""
        <div class="footer">
            Generated by Inactive Mailbox Manager | {datetime.now().year}
        </div>
    </div>
</body>
</html>"""

        return html


class ReportManager:
    """Unified report manager for all formats.

    Routes report generation to appropriate generator based on format.
    """

    def __init__(
        self,
        session: "SessionManager",
        cost_config: CostConfig | None = None,
    ) -> None:
        """Initialize report manager.

        Args:
            session: Session manager
            cost_config: Optional cost configuration
        """
        self._session = session
        self._excel_generator = ExcelReportGenerator(session, cost_config)
        self._html_generator = HTMLReportGenerator(session, cost_config)
        self._audit = session.audit

        logger.debug("ReportManager initialized")

    def generate_report(
        self,
        output_path: Path,
        config: ReportConfig | None = None,
    ) -> ReportResult:
        """Generate a report in the specified format.

        Args:
            output_path: Output file path
            config: Optional report configuration

        Returns:
            ReportResult with status
        """
        if config is None:
            # Detect format from extension
            ext = output_path.suffix.lower()
            format_map = {
                ".xlsx": ReportFormat.EXCEL,
                ".xls": ReportFormat.EXCEL,
                ".html": ReportFormat.HTML,
                ".htm": ReportFormat.HTML,
                ".csv": ReportFormat.CSV,
                ".json": ReportFormat.JSON,
            }
            report_format = format_map.get(ext, ReportFormat.EXCEL)
            config = ReportConfig(format=report_format)

        logger.info(f"Generating {config.format.value} report to {output_path}")

        # Route to appropriate generator
        if config.format == ReportFormat.EXCEL:
            result = self._excel_generator.generate_full_report(output_path, config)
        elif config.format == ReportFormat.HTML:
            result = self._html_generator.generate_report(output_path, config)
        elif config.format == ReportFormat.CSV:
            # Use existing export service
            from src.core.export_service import ExportService
            export_service = ExportService(self._session)
            mailboxes = self._session.db.get_all_mailboxes()
            count = export_service.export_to_csv(mailboxes, output_path)
            result = ReportResult(
                success=True,
                file_path=output_path,
                metadata=ReportMetadata(
                    title="CSV Export",
                    report_type=ReportType.FULL_INVENTORY,
                    generated_at=datetime.now(),
                    generated_by="System",
                    record_count=count,
                ),
            )
        elif config.format == ReportFormat.JSON:
            from src.core.export_service import ExportService
            export_service = ExportService(self._session)
            mailboxes = self._session.db.get_all_mailboxes()
            count = export_service.export_to_json(mailboxes, output_path)
            result = ReportResult(
                success=True,
                file_path=output_path,
                metadata=ReportMetadata(
                    title="JSON Export",
                    report_type=ReportType.FULL_INVENTORY,
                    generated_at=datetime.now(),
                    generated_by="System",
                    record_count=count,
                ),
            )
        else:
            result = ReportResult(
                success=False,
                error=f"Unsupported format: {config.format.value}",
            )

        # Log to audit
        if result.success:
            self._audit.log_export(
                export_path=output_path,
                format=config.format.value,
                record_count=result.metadata.record_count if result.metadata else 0,
            )

        return result

    def get_available_formats(self) -> list[ReportFormat]:
        """Get list of available report formats.

        Returns:
            List of supported formats
        """
        formats = [ReportFormat.CSV, ReportFormat.JSON, ReportFormat.HTML]

        # Check for openpyxl
        try:
            import openpyxl
            formats.insert(0, ReportFormat.EXCEL)
        except ImportError:
            pass

        return formats
